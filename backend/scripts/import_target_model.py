#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

CURRENT_DIR = Path(__file__).resolve().parent
BACKEND_ROOT = CURRENT_DIR.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.append(str(BACKEND_ROOT))

from supabase_key_guard import require_backend_supabase_key  # noqa: E402


@dataclass(frozen=True)
class TargetBucket:
    model_id: str
    portfolio_scope: str
    bucket_key: str
    bucket_label: str
    parent_bucket_key: str | None
    target_weight_pct: float
    lower_band_pct: float | None
    upper_band_pct: float | None
    source_sheet: str
    source_row: int


@dataclass(frozen=True)
class TargetEnvelopeLine:
    model_id: str
    portfolio_scope: str
    envelope: str
    ticker: str | None
    isin: str | None
    instrument: str | None
    asset_class: str | None
    region: str | None
    currency: str | None
    target_weight_pct: float | None
    target_value_eur: float | None
    notes: str | None
    source_sheet: str
    source_row: int


@dataclass(frozen=True)
class TargetAuditHolding:
    model_id: str
    portfolio_scope: str
    envelope: str
    ticker: str | None
    isin: str | None
    instrument: str | None
    asset_class: str | None
    region: str | None
    currency: str | None
    market_value_eur: float | None
    quantity: float | None
    notes: str | None
    source_sheet: str
    source_row: int


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_upper(value: Any) -> str | None:
    text = _clean_text(value)
    return text.upper().replace(" ", "") if text else None


def _read_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        if value.startswith("="):
            return None
        value = value.strip().replace("%", "").replace(",", ".")
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed


def _weight_pct(value: Any) -> float | None:
    parsed = _read_float(value)
    if parsed is None:
        return None
    return parsed * 100 if 0 <= parsed <= 1 else parsed


def _bucket_key(label: str | None) -> str:
    text = (label or "").lower()
    text = text.replace("é", "e").replace("è", "e").replace("à", "a")
    if "cash" in text or "obligation" in text:
        return "cash_bonds"
    if "or" == text.strip() or "gold" in text:
        return "gold"
    if "japon" in text or "japan" in text:
        return "actions_japan"
    if "pac" in text:
        return "actions_pacific_ex_japan"
    if "emerg" in text or "em " in f"{text} ":
        return "actions_emerging"
    if "europe" in text:
        return "actions_europe"
    if "us" in text or "s&p" in text or "sp 500" in text:
        return "actions_us"
    slug = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return slug or "unclassified"


def _target_model_id(kind: str) -> str:
    return f"target_model:{kind.lower()}:active"


def _header_map(values: tuple[Any, ...]) -> dict[str, int]:
    return {
        str(value).strip(): index
        for index, value in enumerate(values)
        if value is not None and str(value).strip()
    }


def _cell(row: tuple[Any, ...], headers: dict[str, int], name: str) -> Any:
    index = headers.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


def parse_personal_model(path: str | Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    model_id = _target_model_id("perso")
    buckets: list[TargetBucket] = []
    envelope_lines: list[TargetEnvelopeLine] = []
    audit_holdings: list[TargetAuditHolding] = []
    warnings: list[str] = []

    strategic = workbook["Strategic_Target_Perso"]
    for row_number, row in enumerate(strategic.iter_rows(min_row=2, values_only=True), start=2):
        label = _clean_text(row[0] if len(row) > 0 else None)
        target = _weight_pct(row[1] if len(row) > 1 else None)
        if not label or target is None:
            continue
        buckets.append(
            TargetBucket(
                model_id=model_id,
                portfolio_scope="PERSO",
                bucket_key=_bucket_key(label),
                bucket_label=label,
                parent_bucket_key=None,
                target_weight_pct=target,
                lower_band_pct=_weight_pct(row[2] if len(row) > 2 else None),
                upper_band_pct=_weight_pct(row[3] if len(row) > 3 else None),
                source_sheet="Strategic_Target_Perso",
                source_row=row_number,
            )
        )

    envelope = workbook["Envelope_Targets"]
    header_row = next(envelope.iter_rows(min_row=4, max_row=4, values_only=True))
    headers = _header_map(header_row)
    for row_number, row in enumerate(envelope.iter_rows(min_row=5, values_only=True), start=5):
        envelope_name = _clean_text(_cell(row, headers, "Envelope"))
        identifier = _clean_text(_cell(row, headers, "ISIN/Ticker"))
        instrument = _clean_text(_cell(row, headers, "Instrument"))
        target = _weight_pct(_cell(row, headers, "Target % (within envelope)"))
        if not envelope_name:
            continue
        if not identifier or not instrument or target is None:
            warnings.append(f"row {row_number}: optional envelope target skipped for {envelope_name}")
            continue
        envelope_lines.append(
            TargetEnvelopeLine(
                model_id=model_id,
                portfolio_scope="PERSO",
                envelope=envelope_name,
                ticker=None if re.match(r"^[A-Z]{2}[A-Z0-9]{10}$", identifier) else _clean_upper(identifier),
                isin=_clean_upper(identifier) if re.match(r"^[A-Z]{2}[A-Z0-9]{10}$", identifier) else None,
                instrument=instrument,
                asset_class=None,
                region=None,
                currency="EUR",
                target_weight_pct=target,
                target_value_eur=_read_float(_cell(row, headers, "Target Value (EUR)")),
                notes=_clean_text(_cell(row, headers, "Notes")),
                source_sheet="Envelope_Targets",
                source_row=row_number,
            )
        )

    holdings = workbook["Holdings_All"]
    holding_headers = _header_map(next(holdings.iter_rows(min_row=1, max_row=1, values_only=True)))
    for row_number, row in enumerate(holdings.iter_rows(min_row=2, values_only=True), start=2):
        envelope_name = _clean_text(_cell(row, holding_headers, "Envelope"))
        if not envelope_name:
            continue
        identifier = _clean_text(_cell(row, holding_headers, "ISIN/Ticker"))
        audit_holdings.append(
            TargetAuditHolding(
                model_id=model_id,
                portfolio_scope="PERSO",
                envelope=envelope_name,
                ticker=None if identifier and re.match(r"^[A-Z]{2}[A-Z0-9]{10}$", identifier) else _clean_upper(identifier),
                isin=_clean_upper(identifier) if identifier and re.match(r"^[A-Z]{2}[A-Z0-9]{10}$", identifier) else None,
                instrument=_clean_text(_cell(row, holding_headers, "Instrument")),
                asset_class=_clean_text(_cell(row, holding_headers, "Asset_Class")),
                region=_clean_text(_cell(row, holding_headers, "Region")),
                currency=_clean_upper(_cell(row, holding_headers, "Currency")),
                market_value_eur=_read_float(_cell(row, holding_headers, "Market_Value_EUR")),
                quantity=None,
                notes="audit only: current official source remains broker snapshots",
                source_sheet="Holdings_All",
                source_row=row_number,
            )
        )

    return _build_report(
        kind="perso",
        source_file=Path(path).name,
        model_id=model_id,
        model_name="Personal strategic and envelope target",
        buckets=buckets,
        envelope_lines=envelope_lines,
        audit_holdings=audit_holdings,
        warnings=warnings,
    )


def parse_pro_model(path: str | Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    model_id = _target_model_id("pro")
    calc = workbook["Calcul_allocation_cible"]
    gold_weight = _weight_pct(calc["E4"].value) or 10.0
    equity_weight = 100.0 - gold_weight
    regional_rows = [
        ("Actions US", "actions_us", "E8"),
        ("Actions Europe", "actions_europe", "E9"),
        ("Actions Japon", "actions_japan", "E10"),
        ("Actions Pacifique ex-JP", "actions_pacific_ex_japan", "E11"),
        ("Actions Emergents", "actions_emerging", "E12"),
    ]

    buckets: list[TargetBucket] = []
    for label, key, cell in regional_rows:
        regional_weight = _weight_pct(calc[cell].value) or 0.0
        buckets.append(
            TargetBucket(
                model_id=model_id,
                portfolio_scope="PRO",
                bucket_key=key,
                bucket_label=label,
                parent_bucket_key="actions",
                target_weight_pct=round(equity_weight * regional_weight / 100.0, 6),
                lower_band_pct=None,
                upper_band_pct=None,
                source_sheet="Calcul_allocation_cible",
                source_row=int(cell[1:]),
            )
        )
    buckets.append(
        TargetBucket(
            model_id=model_id,
            portfolio_scope="PRO",
            bucket_key="gold",
            bucket_label="Or",
            parent_bucket_key=None,
            target_weight_pct=gold_weight,
            lower_band_pct=None,
            upper_band_pct=None,
            source_sheet="Calcul_allocation_cible",
            source_row=4,
        )
    )

    target_by_key = {bucket.bucket_key: bucket.target_weight_pct for bucket in buckets}
    envelope_lines: list[TargetEnvelopeLine] = []
    for row_number, row in enumerate(calc.iter_rows(min_row=16, max_row=21, values_only=True), start=16):
        label = _clean_text(row[0] if len(row) > 0 else None)
        if not label:
            continue
        key = _bucket_key(label)
        envelope_lines.append(
            TargetEnvelopeLine(
                model_id=model_id,
                portfolio_scope="PRO",
                envelope="IBKR_Core",
                ticker=_clean_upper(row[1] if len(row) > 1 else None),
                isin=_clean_upper(row[2] if len(row) > 2 else None),
                instrument=label,
                asset_class=label,
                region=label.replace("Actions ", "").replace("Or", "Gold"),
                currency="EUR",
                target_weight_pct=target_by_key.get(key),
                target_value_eur=None,
                notes="PRO target authority: Calcul_allocation_cible",
                source_sheet="Calcul_allocation_cible",
                source_row=row_number,
            )
        )

    audit_holdings: list[TargetAuditHolding] = []
    ibkr = workbook["IBKR_Positions"]
    ibkr_headers = _header_map(next(ibkr.iter_rows(min_row=5, max_row=5, values_only=True)))
    for row_number, row in enumerate(ibkr.iter_rows(min_row=6, values_only=True), start=6):
        ticker = _clean_upper(_cell(row, ibkr_headers, "Symbol"))
        if not ticker or ticker == "TOTAL":
            continue
        audit_holdings.append(
            TargetAuditHolding(
                model_id=model_id,
                portfolio_scope="PRO",
                envelope="IBKR",
                ticker=ticker,
                isin=None,
                instrument=_clean_text(_cell(row, ibkr_headers, "Description")),
                asset_class=None,
                region=None,
                currency=_clean_upper(_cell(row, ibkr_headers, "Currency")),
                market_value_eur=_read_float(_cell(row, ibkr_headers, "Market Value (EUR)")),
                quantity=_read_float(_cell(row, ibkr_headers, "Quantity")),
                notes="audit only: current official source remains broker snapshots",
                source_sheet="IBKR_Positions",
                source_row=row_number,
            )
        )

    alpheys = workbook["ALPHEYS"]
    alpheys_headers = _header_map(next(alpheys.iter_rows(min_row=5, max_row=5, values_only=True)))
    for row_number, row in enumerate(alpheys.iter_rows(min_row=6, values_only=True), start=6):
        instrument = _clean_text(_cell(row, alpheys_headers, "Instrument"))
        if not instrument or instrument.startswith("TOTAL"):
            continue
        audit_holdings.append(
            TargetAuditHolding(
                model_id=model_id,
                portfolio_scope="PRO",
                envelope="ALPHEYS",
                ticker=None,
                isin=_clean_upper(_cell(row, alpheys_headers, "ISIN")),
                instrument=instrument,
                asset_class="Produit structuré",
                region=None,
                currency="EUR",
                market_value_eur=_read_float(_cell(row, alpheys_headers, "Market Value")),
                quantity=_read_float(_cell(row, alpheys_headers, "Qty")),
                notes="structured product: manual/statement source required for refresh",
                source_sheet="ALPHEYS",
                source_row=row_number,
            )
        )

    cash_buffer = _read_float(workbook["Portefeuille_cible"]["B4"].value)
    return _build_report(
        kind="pro",
        source_file=Path(path).name,
        model_id=model_id,
        model_name="Professional core allocation target",
        buckets=buckets,
        envelope_lines=envelope_lines,
        audit_holdings=audit_holdings,
        warnings=[],
        extra_report={
            "target_authority": "Calcul_allocation_cible",
            "gold_weight_pct": gold_weight,
            "equity_weight_pct": equity_weight,
            "cash_buffer_target_eur": cash_buffer,
        },
    )


def _build_report(
    *,
    kind: str,
    source_file: str,
    model_id: str,
    model_name: str,
    buckets: list[TargetBucket],
    envelope_lines: list[TargetEnvelopeLine],
    audit_holdings: list[TargetAuditHolding],
    warnings: list[str],
    extra_report: dict[str, Any] | None = None,
) -> dict[str, Any]:
    total = round(sum(bucket.target_weight_pct for bucket in buckets), 6)
    rejected: list[dict[str, Any]] = []
    if abs(total - 100.0) > 0.05:
        rejected.append({"reason": f"target bucket total must equal 100% ±0.05 ({total:.4f}%)"})
    report_json = {
        "warnings": warnings,
        "rejected": rejected,
        "bucket_count": len(buckets),
        "envelope_line_count": len(envelope_lines),
        "audit_holding_count": len(audit_holdings),
        **(extra_report or {}),
    }
    return {
        "ok": not rejected,
        "kind": kind,
        "portfolio_scope": kind.upper(),
        "model_id": model_id,
        "model_name": model_name,
        "source_file": source_file,
        "target_total_pct": total,
        "buckets": buckets,
        "envelope_lines": envelope_lines,
        "audit_holdings": audit_holdings,
        "warnings": warnings,
        "rejected": rejected,
        "report_json": report_json,
    }


def parse_target_model(path: str | Path, *, kind: str) -> dict[str, Any]:
    if kind == "perso":
        return parse_personal_model(path)
    if kind == "pro":
        return parse_pro_model(path)
    raise RuntimeError("Unsupported --kind. Expected perso or pro.")


def _build_supabase_client() -> Any:
    from supabase import create_client

    url = os.environ.get("SUPABASE_URL")
    if not url:
        raise RuntimeError("SUPABASE_URL is required for --apply")
    return create_client(url, require_backend_supabase_key(os.environ))


def _payload(row: Any) -> dict[str, Any]:
    payload = asdict(row)
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    return payload


def _chunks(rows: list[dict[str, Any]], size: int = 500):
    for index in range(0, len(rows), size):
        yield rows[index:index + size]


def apply_target_model(report: dict[str, Any], *, supabase_client: Any) -> dict[str, Any]:
    now = datetime.now(timezone.utc).isoformat()
    model_payload = {
        "id": report["model_id"],
        "portfolio_scope": report["portfolio_scope"],
        "model_name": report["model_name"],
        "source_file": report["source_file"],
        "source_kind": report["kind"],
        "as_of_date": None,
        "is_active": True,
        "target_total_pct": report["target_total_pct"],
        "status": "READY" if report["ok"] else "INVALID",
        "report_json": report["report_json"],
        "updated_at": now,
    }
    supabase_client.table("target_models").upsert(model_payload, on_conflict="id").execute()

    for table in ("target_buckets", "target_envelope_lines", "target_model_audit_holdings"):
        supabase_client.table(table).delete().eq("model_id", report["model_id"]).execute()

    bucket_payloads = [_payload(row) for row in report["buckets"]]
    envelope_payloads = [_payload(row) for row in report["envelope_lines"]]
    audit_payloads = [_payload(row) for row in report["audit_holdings"]]

    for chunk in _chunks(bucket_payloads):
        supabase_client.table("target_buckets").insert(chunk).execute()
    for chunk in _chunks(envelope_payloads):
        supabase_client.table("target_envelope_lines").insert(chunk).execute()
    for chunk in _chunks(audit_payloads):
        supabase_client.table("target_model_audit_holdings").insert(chunk).execute()

    return {
        "model_upserted": report["model_id"],
        "buckets_inserted": len(bucket_payloads),
        "envelope_lines_inserted": len(envelope_payloads),
        "audit_holdings_inserted": len(audit_payloads),
    }


def run_import(
    input_path: str | Path,
    *,
    kind: str,
    dry_run: bool = True,
    supabase_client: Any | None = None,
) -> dict[str, Any]:
    report = parse_target_model(input_path, kind=kind)
    write_report = {
        "model_upserted": None,
        "buckets_inserted": 0,
        "envelope_lines_inserted": 0,
        "audit_holdings_inserted": 0,
    }
    if not dry_run:
        if supabase_client is None:
            raise RuntimeError("A Supabase client is required when dry_run=False")
        write_report = apply_target_model(report, supabase_client=supabase_client)

    return {
        "ok": report["ok"],
        "dry_run": dry_run,
        "kind": report["kind"],
        "portfolio_scope": report["portfolio_scope"],
        "model_id": report["model_id"],
        "source_file": report["source_file"],
        "target_total_pct": report["target_total_pct"],
        "bucket_count": len(report["buckets"]),
        "envelope_line_count": len(report["envelope_lines"]),
        "audit_holding_count": len(report["audit_holdings"]),
        "buckets": [asdict(row) for row in report["buckets"]],
        "envelope_lines": [asdict(row) for row in report["envelope_lines"]],
        "warnings": report["warnings"],
        "rejected": report["rejected"],
        "write": write_report,
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import two-level portfolio target models from curated Excel files")
    parser.add_argument("--kind", required=True, choices=["perso", "pro"], help="Target model workbook kind")
    parser.add_argument("--file", required=True, help="Target model .xlsx file")
    parser.add_argument("--dry-run", action="store_true", help="Validate and report without writing Supabase")
    parser.add_argument("--apply", action="store_true", help="Write target model to Supabase")
    args = parser.parse_args()
    if args.dry_run and args.apply:
        parser.error("--dry-run and --apply are mutually exclusive")
    if not args.dry_run and not args.apply:
        args.dry_run = True
    return args


def main() -> int:
    args = _parse_args()
    try:
        client = _build_supabase_client() if args.apply else None
        report = run_import(args.file, kind=args.kind, dry_run=args.dry_run, supabase_client=client)
    except Exception as exc:
        report = {
            "ok": False,
            "dry_run": args.dry_run,
            "source_file": Path(args.file).name,
            "error": str(exc),
        }
    print(json.dumps(report, indent=2, ensure_ascii=False, default=str))
    return 0 if report.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
