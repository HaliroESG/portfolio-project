#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

CURRENT_DIR = Path(__file__).resolve().parent
BACKEND_ROOT = CURRENT_DIR.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.append(str(BACKEND_ROOT))

from supabase_key_guard import require_backend_supabase_key  # noqa: E402


TARGET_TOTAL_TOLERANCE = 0.05

HEADER_ALIASES = {
    "portfolio_id": {"portfolioid", "portfolio", "portfolio_id", "portfolio id"},
    "ticker": {"ticker", "symbol", "symbole"},
    "isin": {"isin", "codeisin", "code isin"},
    "name": {"name", "nom", "asset", "instrument", "security"},
    "asset_class": {"assetclass", "asset_class", "classeactif", "classe d actif", "class"},
    "currency": {"currency", "devise", "ccy"},
    "target_weight_pct": {
        "targetweightpct",
        "target_weight_pct",
        "targetpct",
        "target_pct",
        "weightpct",
        "weight_pct",
        "poidscible",
        "allocationcible",
    },
    "notes": {"notes", "note", "comment", "comments", "commentaire"},
}


@dataclass(frozen=True)
class RawTargetRow:
    row_number: int
    portfolio_id: str
    ticker: str | None
    isin: str | None
    name: str | None
    asset_class: str | None
    currency: str | None
    target_weight_pct: float
    notes: str | None


@dataclass(frozen=True)
class TargetAllocationRow:
    row_number: int
    portfolio_id: str
    ticker: str
    isin: str | None
    name: str | None
    asset_class: str | None
    currency: str | None
    target_weight_pct: float
    notes: str | None


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("%", "pct")
    return re.sub(r"[^a-z0-9_]+", "", text)


def _canonical_header(value: Any) -> str | None:
    normalized = _normalize_header(value)
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized == _normalize_header(canonical) or normalized in {_normalize_header(alias) for alias in aliases}:
            return canonical
    return None


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_upper(value: Any) -> str | None:
    text = _clean_text(value)
    return text.upper().replace(" ", "") if text else None


def _read_weight(value: Any, row_number: int) -> float:
    if value is None or value == "":
        raise ValueError(f"row {row_number}: target_weight_pct is required")
    if isinstance(value, str):
        value = value.strip().replace("%", "").replace(",", ".")
    try:
        weight = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"row {row_number}: target_weight_pct is not numeric") from exc
    if weight < 0 or weight > 100:
        raise ValueError(f"row {row_number}: target_weight_pct must be between 0 and 100")
    return weight


def _row_payload(headers: dict[int, str], values: tuple[Any, ...]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for index, canonical in headers.items():
        payload[canonical] = values[index] if index < len(values) else None
    return payload


def parse_target_excel(path: str | Path, *, sheet: str | None = None) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {"rows_read": 0, "accepted": [], "rejected": [{"row": 1, "reason": "empty workbook"}]}

    headers = {
        index: canonical
        for index, value in enumerate(header_row)
        for canonical in [_canonical_header(value)]
        if canonical is not None
    }
    missing_headers = {"portfolio_id", "target_weight_pct"} - set(headers.values())
    if missing_headers:
        return {
            "rows_read": 0,
            "accepted": [],
            "rejected": [{"row": 1, "reason": f"missing headers: {', '.join(sorted(missing_headers))}"}],
        }
    if "ticker" not in set(headers.values()) and "isin" not in set(headers.values()):
        return {
            "rows_read": 0,
            "accepted": [],
            "rejected": [{"row": 1, "reason": "missing ticker or isin header"}],
        }

    accepted: list[RawTargetRow] = []
    rejected: list[dict[str, Any]] = []
    rows_read = 0
    seen: set[tuple[str, str]] = set()

    for row_number, values in enumerate(rows_iter, start=2):
        if not any(value not in (None, "") for value in values):
            continue
        rows_read += 1
        payload = _row_payload(headers, values)
        try:
            portfolio_id = _clean_text(payload.get("portfolio_id")) or ""
            ticker = _clean_upper(payload.get("ticker"))
            isin = _clean_upper(payload.get("isin"))
            if not portfolio_id:
                raise ValueError(f"row {row_number}: portfolio_id is required")
            if not ticker and not isin:
                raise ValueError(f"row {row_number}: ticker or isin is required")
            dedupe_key = (portfolio_id, ticker or f"ISIN:{isin}")
            if dedupe_key in seen:
                raise ValueError(f"row {row_number}: duplicate target for {portfolio_id}/{dedupe_key[1]}")
            seen.add(dedupe_key)
            accepted.append(
                RawTargetRow(
                    row_number=row_number,
                    portfolio_id=portfolio_id,
                    ticker=ticker,
                    isin=isin,
                    name=_clean_text(payload.get("name")),
                    asset_class=_clean_text(payload.get("asset_class")),
                    currency=_clean_upper(payload.get("currency")),
                    target_weight_pct=_read_weight(payload.get("target_weight_pct"), row_number),
                    notes=_clean_text(payload.get("notes")),
                )
            )
        except ValueError as exc:
            rejected.append({"row": row_number, "reason": str(exc)})

    return {
        "rows_read": rows_read,
        "accepted": accepted,
        "rejected": rejected,
    }


def validate_target_totals(rows: list[TargetAllocationRow], *, tolerance: float = TARGET_TOTAL_TOLERANCE) -> dict[str, float]:
    totals: dict[str, float] = {}
    for row in rows:
        totals[row.portfolio_id] = totals.get(row.portfolio_id, 0.0) + row.target_weight_pct

    invalid = {
        portfolio_id: total
        for portfolio_id, total in totals.items()
        if abs(total - 100.0) > tolerance
    }
    if invalid:
        formatted = ", ".join(f"{portfolio_id}={total:.2f}%" for portfolio_id, total in sorted(invalid.items()))
        raise ValueError(f"Target allocations must sum to 100% ±{tolerance:.2f} per portfolio ({formatted}).")
    return totals


def _build_supabase_client() -> Any:
    from supabase import create_client

    url = os.environ.get("SUPABASE_URL")
    if not url:
        raise RuntimeError("SUPABASE_URL is required for --apply")
    return create_client(url, require_backend_supabase_key(os.environ))


def _response_data(response: Any) -> list[dict[str, Any]]:
    data = getattr(response, "data", response.get("data") if isinstance(response, dict) else None)
    return data or []


def resolve_ticker_from_isin(supabase_client: Any, isin: str) -> str | None:
    response = (
        supabase_client
        .table("instrument_identifier_map")
        .select("ticker")
        .eq("isin", isin)
        .limit(1)
        .execute()
    )
    rows = _response_data(response)
    if not rows:
        return None
    ticker = _clean_upper(rows[0].get("ticker"))
    return ticker


def resolve_target_rows(
    raw_rows: list[RawTargetRow],
    *,
    supabase_client: Any | None = None,
) -> tuple[list[TargetAllocationRow], list[dict[str, Any]], list[str]]:
    resolved: list[TargetAllocationRow] = []
    rejected: list[dict[str, Any]] = []
    warnings: list[str] = []

    for row in raw_rows:
        ticker = row.ticker
        if not ticker and row.isin and supabase_client is not None:
            try:
                ticker = resolve_ticker_from_isin(supabase_client, row.isin)
            except Exception as exc:
                warnings.append(f"row {row.row_number}: ISIN resolution failed for {row.isin}: {exc}")
        if not ticker:
            rejected.append({
                "row": row.row_number,
                "reason": f"ticker missing and ISIN {row.isin or ''} could not be resolved",
            })
            continue
        resolved.append(
            TargetAllocationRow(
                row_number=row.row_number,
                portfolio_id=row.portfolio_id,
                ticker=ticker,
                isin=row.isin,
                name=row.name,
                asset_class=row.asset_class,
                currency=row.currency,
                target_weight_pct=row.target_weight_pct,
                notes=row.notes,
            )
        )

    return resolved, rejected, warnings


def _position_update_payload(row: TargetAllocationRow, *, source_file: str | None = None) -> dict[str, Any]:
    now = datetime.now(timezone.utc).isoformat()
    payload: dict[str, Any] = {
        "target_weight_pct": row.target_weight_pct,
        "target_source": "excel",
        "target_updated_at": now,
        "updated_at": now,
    }
    if source_file:
        payload["target_source_file"] = source_file
    if row.name:
        payload["name"] = row.name
    if row.asset_class:
        payload["instrument_type"] = row.asset_class
    if row.currency:
        payload["currency"] = row.currency
    if row.isin:
        payload["isin"] = row.isin
    if row.notes:
        payload["target_notes"] = row.notes
    return payload


def _position_insert_payload(row: TargetAllocationRow, *, source_file: str | None = None) -> dict[str, Any]:
    payload = _position_update_payload(row, source_file=source_file)
    payload.update({
        "portfolio_id": row.portfolio_id,
        "ticker": row.ticker,
        "quantity_current": 0,
    })
    return payload


def _existing_position(supabase_client: Any, row: TargetAllocationRow) -> bool:
    response = (
        supabase_client
        .table("portfolio_positions")
        .select("id")
        .eq("portfolio_id", row.portfolio_id)
        .eq("ticker", row.ticker)
        .limit(1)
        .execute()
    )
    return bool(_response_data(response))


def import_target_allocations(
    rows: list[TargetAllocationRow],
    *,
    dry_run: bool,
    supabase_client: Any | None = None,
    source_file: str | None = None,
) -> dict[str, Any]:
    updated: list[dict[str, Any]] = []
    inserted: list[dict[str, Any]] = []

    if dry_run:
        for row in rows:
            inserted.append({
                "portfolio_id": row.portfolio_id,
                "ticker": row.ticker,
                "target_weight_pct": row.target_weight_pct,
                "dry_run": True,
                "action": "update_or_insert",
            })
        return {"updated": updated, "inserted": inserted}

    if supabase_client is None:
        raise RuntimeError("A Supabase client is required when dry_run=False")

    for row in rows:
        if _existing_position(supabase_client, row):
            payload = _position_update_payload(row, source_file=source_file)
            (
                supabase_client
                .table("portfolio_positions")
                .update(payload)
                .eq("portfolio_id", row.portfolio_id)
                .eq("ticker", row.ticker)
                .execute()
            )
            updated.append({
                "portfolio_id": row.portfolio_id,
                "ticker": row.ticker,
                "target_weight_pct": row.target_weight_pct,
            })
            continue

        payload = _position_insert_payload(row, source_file=source_file)
        supabase_client.table("portfolio_positions").insert(payload).execute()
        inserted.append({
            "portfolio_id": row.portfolio_id,
            "ticker": row.ticker,
            "target_weight_pct": row.target_weight_pct,
        })

    return {"updated": updated, "inserted": inserted}


def run_import(
    input_path: str | Path,
    *,
    sheet: str | None = None,
    dry_run: bool = True,
    supabase_client: Any | None = None,
) -> dict[str, Any]:
    parse_report = parse_target_excel(input_path, sheet=sheet)
    raw_rows: list[RawTargetRow] = parse_report["accepted"]
    rejected = list(parse_report["rejected"])
    warnings: list[str] = []

    resolved_rows, resolution_rejected, resolution_warnings = resolve_target_rows(
        raw_rows,
        supabase_client=supabase_client,
    )
    rejected.extend(resolution_rejected)
    warnings.extend(resolution_warnings)

    portfolio_totals: dict[str, float] = {}
    write_report = {"updated": [], "inserted": []}
    ok = False
    error: str | None = None

    try:
        portfolio_totals = validate_target_totals(resolved_rows)
        write_report = import_target_allocations(
            resolved_rows,
            dry_run=dry_run,
            supabase_client=supabase_client,
            source_file=Path(input_path).name,
        )
        ok = len(rejected) == 0
    except Exception as exc:
        error = str(exc)

    return {
        "ok": ok,
        "dry_run": dry_run,
        "source_file": Path(input_path).name,
        "rows_read": parse_report["rows_read"],
        "rows_accepted": len(resolved_rows),
        "rows_rejected": len(rejected),
        "portfolio_totals": portfolio_totals,
        "updated": write_report["updated"],
        "inserted": write_report["inserted"],
        "rejected": rejected,
        "warnings": warnings,
        "error": error,
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import target portfolio allocation from a simple Excel file")
    parser.add_argument("--file", required=True, help="Target allocation .xlsx file")
    parser.add_argument("--sheet", default=None, help="Optional worksheet name")
    parser.add_argument("--dry-run", action="store_true", help="Validate and report without writing Supabase")
    parser.add_argument("--apply", action="store_true", help="Write target allocations to Supabase")
    args = parser.parse_args()
    if args.dry_run and args.apply:
        parser.error("--dry-run and --apply are mutually exclusive")
    if not args.dry_run and not args.apply:
        args.dry_run = True
    return args


def main() -> int:
    args = _parse_args()
    try:
        client = _build_supabase_client() if args.apply else None
        report = run_import(
            args.file,
            sheet=args.sheet,
            dry_run=args.dry_run,
            supabase_client=client,
        )
    except Exception as exc:
        report = {
            "ok": False,
            "dry_run": args.dry_run,
            "source_file": Path(args.file).name,
            "error": str(exc),
        }
    print(json.dumps(report, indent=2, sort_keys=True))
    return 0 if report.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
