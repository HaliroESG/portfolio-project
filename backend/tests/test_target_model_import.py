from __future__ import annotations

from openpyxl import Workbook

from scripts.import_target_model import parse_target_model, run_import


class _Response:
    def __init__(self, data=None):
        self.data = data or []


class _Table:
    def __init__(self, client, name: str):
        self.client = client
        self.name = name
        self.operation = None
        self.payload = None
        self.filters = {}
        self.on_conflict = None

    def upsert(self, payload, on_conflict):
        self.operation = "upsert"
        self.payload = payload
        self.on_conflict = on_conflict
        return self

    def delete(self):
        self.operation = "delete"
        return self

    def insert(self, payload):
        self.operation = "insert"
        self.payload = payload
        return self

    def eq(self, column, value):
        self.filters[column] = value
        return self

    def execute(self):
        rows = self.client.rows.setdefault(self.name, [])
        if self.operation == "upsert":
            key = self.on_conflict
            existing = next((row for row in rows if row.get(key) == self.payload.get(key)), None)
            if existing:
                existing.update(self.payload)
                return _Response([existing])
            rows.append(self.payload.copy())
            return _Response([self.payload.copy()])
        if self.operation == "delete":
            self.client.rows[self.name] = [
                row for row in rows
                if any(row.get(column) != value for column, value in self.filters.items())
            ]
            return _Response([])
        if self.operation == "insert":
            payloads = self.payload if isinstance(self.payload, list) else [self.payload]
            rows.extend(payload.copy() for payload in payloads)
            return _Response(payloads)
        return _Response([])


class _Supabase:
    def __init__(self):
        self.rows = {}

    def table(self, name):
        return _Table(self, name)


def _write_personal(path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Strategic_Target_Perso"
    ws.append(["Bucket", "Target %", "Lower Band %", "Upper Band %"])
    ws.append(["Actions US", 0.42, 0.35, 0.55])
    ws.append(["Actions Europe", 0.23, 0.15, 0.30])
    ws.append(["Actions Japon", 0.07, 0.03, 0.12])
    ws.append(["Actions Emergents", 0.13, 0.05, 0.20])
    ws.append(["Or", 0.08, 0.04, 0.12])
    ws.append(["Obligations / Cash", 0.07, 0.03, 0.15])

    ws = workbook.create_sheet("Envelope_Targets")
    ws.append(["Envelope-level targets"])
    ws.append([])
    ws.append([])
    ws.append(["Envelope", "ISIN/Ticker", "Instrument", "Target % (within envelope)", "Target Value (EUR)", "Notes"])
    ws.append(["Cardif_Lucya_PostArb", "LU0496786574", "Amundi Core S&P 500 Swap ETF", 0.56, 5652.17, "core"])
    ws.append(["Fortuneo_CTO", None, None, None, None, "Optional"])

    ws = workbook.create_sheet("Holdings_All")
    ws.append(["Envelope", "ISIN/Ticker", "Instrument", "Asset_Class", "Region", "Currency", "Market_Value_EUR"])
    ws.append(["Fortuneo_PEA", "AI", "AIR LIQUIDE", "Equity", "Europe", "EUR", 8098.8])
    ws.append(["Cardif_Lucya_PostArb", "LU0496786574", "Amundi Core S&P 500 Swap ETF", "Equity", "US", "EUR", 5652.17])
    workbook.save(path)


def _write_pro(path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Calcul_allocation_cible"
    for _ in range(25):
        ws.append([None] * 10)
    ws["E4"] = 0.1
    ws["E8"] = 0.45
    ws["E9"] = 0.2
    ws["E10"] = 0.1
    ws["E11"] = 0.05
    ws["E12"] = 0.2
    target_rows = [
        ("Actions US", "CSPX", "IE00B5BMR087"),
        ("Actions Europe", "IMAE", "IE00B4K48X80"),
        ("Actions Japon", "CJPU", "IE00B4L5YX21"),
        ("Actions Pac ex-JP", "CPXJ", "IE00B52MJY50"),
        ("Actions EM IMI", "EIMI", "IE00BKM4GZ66"),
        ("Or", "GOLD", "FR0013416716"),
    ]
    for index, values in enumerate(target_rows, start=16):
        ws.cell(row=index, column=1, value=values[0])
        ws.cell(row=index, column=2, value=values[1])
        ws.cell(row=index, column=3, value=values[2])

    ws = workbook.create_sheet("Portefeuille_cible")
    ws["B4"] = 100000

    ws = workbook.create_sheet("IBKR_Positions")
    for _ in range(4):
        ws.append([])
    ws.append(["Symbol", "Description", "Currency", "Quantity", "Price", "Market Value (ccy)", "FX to EUR", "Market Value (EUR)"])
    ws.append(["CSPX", "ISHARES CORE S&P 500", "USD", 2, 806, 1612, 0.86, 1386])
    ws.append(["TOTAL", None, None, None, None, None, None, 1386])

    ws = workbook.create_sheet("ALPHEYS")
    for _ in range(4):
        ws.append([])
    ws.append(["Instrument", "ISIN", "Qty", "Nominal", "Prix (EUR)", "Market Value"])
    ws.append(["Athena", "XS3103296227", 25, 1000, 682.7, 17067.5])
    ws.append(["TOTAL titres", None, None, None, None, 17067.5])
    workbook.save(path)


def test_personal_target_model_reads_global_and_envelope_targets(tmp_path):
    source = tmp_path / "personal.xlsx"
    _write_personal(source)

    report = parse_target_model(source, kind="perso")

    assert report["ok"] is True
    assert report["target_total_pct"] == 100
    assert len(report["buckets"]) == 6
    assert len(report["envelope_lines"]) == 1
    assert report["envelope_lines"][0].envelope == "Cardif_Lucya_PostArb"
    assert report["audit_holdings"][0].notes.startswith("audit only")


def test_pro_target_model_uses_calculation_sheet_authority(tmp_path):
    source = tmp_path / "pro.xlsx"
    _write_pro(source)

    report = parse_target_model(source, kind="pro")
    buckets = {row.bucket_key: row.target_weight_pct for row in report["buckets"]}

    assert report["ok"] is True
    assert buckets["gold"] == 10
    assert buckets["actions_us"] == 40.5
    assert buckets["actions_europe"] == 18
    assert buckets["actions_japan"] == 9
    assert buckets["actions_pacific_ex_japan"] == 4.5
    assert buckets["actions_emerging"] == 18


def test_target_model_apply_replaces_child_rows(tmp_path):
    source = tmp_path / "personal.xlsx"
    _write_personal(source)
    fake = _Supabase()

    report = run_import(source, kind="perso", dry_run=False, supabase_client=fake)

    assert report["ok"] is True
    assert fake.rows["target_models"][0]["id"] == "target_model:perso:active"
    assert len(fake.rows["target_buckets"]) == 6
    assert len(fake.rows["target_envelope_lines"]) == 1
    assert len(fake.rows["target_model_audit_holdings"]) == 2
